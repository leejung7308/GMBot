import discord
from discord.ext import commands, tasks
from discord.ui import Button, View
import random
import asyncio
import json
import os
from datetime import datetime, timedelta, timezone
import openpyxl
from io import BytesIO
import random
import re
import pytz

kst = pytz.timezone('Asia/Seoul')

# 봇 초기화
intents = discord.Intents.default()
intents.messages = True  # 메시지를 읽기 위한 권한
intents.reactions = True  # 반응을 읽기 위한 권한
intents.message_content = True  # 메시지 내용을 읽기 위한 권한
intents.members = True  # 멤버 목록을 읽기 위한 권한
intents.guilds = True  # 서버 목록을 읽기 위한 권한

bot = commands.Bot(command_prefix='!', intents=intents)

CONFIG_FILE = 'config.json'
GUILD_FILE = 'guilds.json'
RANKING_FILE = 'ranking.json'

def save_config(data):
    with open(CONFIG_FILE, 'w') as f:
        json.dump(data, f, indent=4)

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r') as f:
                return json.load(f)
        except json.JSONDecodeError:
            # JSONDecodeError 발생 시 기본값으로 초기화
            return {}
    return {}

def save_guilds(data):
    with open(GUILD_FILE, 'w') as f:
        json.dump(data, f, indent=4)

def load_guilds():
    if os.path.exists(GUILD_FILE):
        try:
            with open(GUILD_FILE, 'r') as f:
                return json.load(f)
        except json.JSONDecodeError:
            print("길드 파일이 손상되었습니다.")
            return {}
    return {}

def save_ranking(data):
    for guild_id in data:
        for user_id, user_data in data[guild_id].items():
            if 'join_time' in user_data:
                if isinstance(user_data['join_time'], datetime):
                    user_data['join_time'] = user_data['join_time'].isoformat()
    with open(RANKING_FILE, 'w') as f:
        json.dump(data, f, indent=4)

def load_ranking():
    if os.path.exists(RANKING_FILE):
        try:
            with open(RANKING_FILE, 'r') as f:
                return json.load(f)
        except json.JSONDecodeError:
            print("랭킹 파일이 손상되었습니다.")
            return {}
    return {}

@bot.event
async def on_ready():
    bot.add_view(PersistentView())
    await bot.change_presence(activity=discord.Game(name="인사"))
    print(f'Logged in as {bot.user}')
    configs = load_config()
    update_rankings.start()

@bot.event
async def on_message(message):
    if message.author == bot.user:
        return

    if message.content.startswith('!'):
        await bot.process_commands(message)

    if message.content.startswith('!') and not message.content.startswith('!청소'):
        await message.delete()
    
    ranking_data = load_ranking()
    user_id = str(message.author.id)
    guild_id = str(message.guild.id)
    if guild_id not in ranking_data:
        ranking_data[guild_id] = {}
    if user_id not in ranking_data[guild_id]:
        ranking_data[guild_id][user_id] = {"message_count": 0, "voice_time": 0}
    ranking_data[guild_id][user_id]["message_count"] += 1

    save_ranking(ranking_data)

@bot.event
async def on_voice_state_update(member, before, after):
    ranking_data = load_ranking()
    guild_id = str(member.guild.id)
    user_id = str(member.id)
    if guild_id not in ranking_data:
        ranking_data[guild_id] = {}
    if user_id not in ranking_data[guild_id]:
        ranking_data[guild_id][user_id] = {"message_count": 0, "voice_time": 0}
    if before.channel is None and after.channel is not None:
        ranking_data[guild_id][user_id]['join_time'] = datetime.now(kst).isoformat()
    elif before.channel is not None and after.channel is None:
        if 'join_time' in ranking_data[guild_id][user_id]:
            join_time_str = ranking_data[guild_id][user_id].pop('join_time')
            join_time = datetime.fromisoformat(join_time_str)
            leave_time = datetime.now(kst)
            duration = leave_time - join_time
            ranking_data[guild_id][user_id]['voice_time'] += duration.total_seconds()

    save_ranking(ranking_data)

@tasks.loop(minutes=1)
async def update_rankings():
    now = datetime.now(kst)
    if now.minute == 0:
        await announce_rankings()

async def announce_rankings():
    now = datetime.now(kst)
    ranking_data = load_ranking()
    configs = load_config()
    if not ranking_data:
        return

    for guild_id in ranking_data:
        for user_id, user_data in ranking_data[guild_id].items():
            if 'join_time' in user_data:
                join_time_str = user_data.pop('join_time')
                join_time = datetime.fromisoformat(join_time_str)
                leave_time = datetime.now(kst)
                duration = leave_time - join_time
                user_data['voice_time'] += duration.total_seconds()
                ranking_data[guild_id][user_id]['join_time'] = leave_time.isoformat()
        sorted_by_message_count = sorted(ranking_data[guild_id].items(), key=lambda x: x[1]['message_count'], reverse=True)
        sorted_by_voice_time = sorted(ranking_data[guild_id].items(), key=lambda x: x[1]['voice_time'], reverse=True)
        guild = bot.get_guild(int(guild_id))
        channel_id = configs[guild_id].get('ranking_channel_id')
        channel = bot.get_channel(channel_id)
        message_id = configs[guild_id].get('ranking_message_id')
        if message_id and channel:
            message = await channel.fetch_message(message_id)
        if not message:
            return
        embed = message.embeds[0]
        embed.title = f"📊랭킹 현황({now.strftime('%Y년 %m월 %d일, %H:%M')})📊"
        embed.description = "랭킹은 매시 00분에 업데이트됩니다."
        embed.set_field_at(0, name=f"⌨️ {guild.name} 키보드워리어 랭킹 ⌨️", value="\n".join([f"**{idx + 1}**. <@{user_id}>: {data['message_count']}개" for idx, (user_id, data) in enumerate(sorted_by_message_count[:5])]), inline=False)
        embed.set_field_at(1, name=f"👻 {guild.name} 지박령 랭킹 👻", value="\n".join([f"**{idx + 1}**. <@{user_id}>: {timedelta(seconds=int(data['voice_time']))}" for idx, (user_id, data) in enumerate(sorted_by_voice_time[:5])]), inline=False)
        save_config(configs)
        save_ranking(ranking_data)
        await message.edit(embed=embed)

class PersistentView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)
        self.add_item(ShowEntireRankingButton(label="전체 랭킹 보기", style=discord.ButtonStyle.primary))
        #self.add_item(RefreshRankingButton(label="랭킹 갱신하기", style=discord.ButtonStyle.secondary))

class ShowEntireRankingButton(discord.ui.Button):
    def __init__(self, label, style):
        super().__init__(label=label, style=style, custom_id="show_entire_ranking_button")
    
    async def callback(self, interaction: discord.Interaction):
        view = self.view
        self.disabled = True
        await interaction.message.edit(view=view)
        await interaction.response.defer()
        ranking_data = load_ranking()
        guild_id = str(interaction.guild.id)
        guild = bot.get_guild(interaction.guild.id)
        sorted_by_message_count = sorted(ranking_data[guild_id].items(), key=lambda x: x[1]['message_count'], reverse=True)
        sorted_by_voice_time = sorted(ranking_data[guild_id].items(), key=lambda x: x[1]['voice_time'], reverse=True)
        embed = discord.Embed(
            title="📊실시간 전체 랭킹📊",
            description="랭킹은 매시 00분에 업데이트됩니다.",
            color=discord.Color.blue()
        )
        embed.add_field(name=f"⌨️ {guild.name} 키보드워리어 랭킹 ⌨️", value="\n".join([f"**{idx + 1}**. <@{user_id}>: {data['message_count']}개" for idx, (user_id, data) in enumerate(sorted_by_message_count)]), inline=False)
        embed.add_field(name=f"👻 {guild.name} 지박령 랭킹 👻", value="\n".join([f"**{idx + 1}**. <@{user_id}>: {timedelta(seconds=int(data['voice_time']))}" for idx, (user_id, data) in enumerate(sorted_by_voice_time)]), inline=False)
        await interaction.user.send(embed=embed)

        self.label = "전체 랭킹 보기"
        self.disabled = False
        await interaction.message.edit(view=view)

'''class RefreshRankingButton(discord.ui.Button):
    def __init__(self, label, style):
        super().__init__(label=label, style=style, custom_id="refresh_ranking_button")
    
    async def callback(self, interaction: discord.Interaction):
        view = self.view
        self.disabled = True
        await interaction.message.edit(view=view)
        await interaction.response.defer()
        await announce_rankings()

        self.label = "랭킹 갱신하기"
        self.disabled = False
        await interaction.message.edit(view=view)'''
        

@bot.event
async def on_member_join(member):
    role = discord.utils.get(member.guild.roles, name="새싹")
    if role:
        try:
            await member.add_roles(role)
        except discord.Forbidden:
            print(f"권한부족: {member}에게 역할을 부여할 수 없습니다.")
        except discord.HTTPException as e:
            print(f"오류 발생: {e}")

@bot.event
async def on_raw_reaction_add(payload):

    if payload.member.bot:
        return
    guild_id = str(payload.guild_id)

    ## 역할 부여
    config = load_config()
    if guild_id in config:
        announcement_message_id = config[guild_id].get('announcement_message_id')
        if announcement_message_id:
            if payload.message_id == announcement_message_id:
                guild = bot.get_guild(payload.guild_id)
                if guild:
                    role = None
                    if str(payload.emoji) == '📝':
                        role = discord.utils.get(guild.roles, name="기획")
                    if str(payload.emoji) == '💻':
                        role = discord.utils.get(guild.roles, name="프로그래밍")
                    if str(payload.emoji) == '🎨':
                        role = discord.utils.get(guild.roles, name="아트")
                    if str(payload.emoji) == '🎵':
                        role = discord.utils.get(guild.roles, name="사운드")
                    if str(payload.emoji) == '✅':
                        role = discord.utils.get(guild.roles, name="GM 일반멤버")
        
                    if role is not None:
                        member = guild.get_member(payload.user_id)
                        if member is not None:
                            try:
                                await member.add_roles(role)
                                await member.remove_roles(discord.utils.get(guild.roles, name="새싹"))
                                if guild_id in config and 'welcome_channel_id' in config[guild_id]:
                                    welcome_channel_id = config[guild_id]['welcome_channel_id']
                                    welcome_channel = bot.get_channel(welcome_channel_id)
                                if welcome_channel:
                                    # 공지 채널에 환영 메시지 전송
                                    embed = discord.Embed(
                                        title="🖐️GameMakers에 오신 것을 환영합니다!🖐️",
                                        description=f"모두 {member.mention}님을 환영해주세요!",
                                        color=discord.Color.blue()
                                    )
                                    await welcome_channel.send(embed=embed)
                            except discord.Forbidden:
                                print(f"권한부족: {member}에게 역할을 부여할 수 없습니다.")
                            except discord.HTTPException as e:
                                print(f"오류 발생: {e}")

    ## 길드 가입
    guilds = load_guilds()
    if guild_id in guilds: #얘는 길드
        for guild_name, guild_info in guilds[guild_id].items():
            if isinstance(guild_info, dict):
                if payload.message_id == guild_info['message_id']:
                    guild = bot.get_guild(payload.guild_id)
                    if guild: #얘는 서버
                        member = guild.get_member(payload.user_id)
                        if member:
                            if str(payload.emoji) == '📝':
                                guild_role = guild.get_role(guild_info['role_id'])
                                if guild_role:
                                    if 'guild_members' not in guild_info:
                                        guild_info['guild_members'] = []
                                    channel = bot.get_channel(payload.channel_id)
                                    message = await channel.fetch_message(payload.message_id)
                                    if member.id in guild_info['guild_members']:
                                        await member.send("이미 가입된 길드입니다.")
                                        await message.remove_reaction(payload.emoji, member)
                                        return
                                    elif member.id == guild_info['guild_leader_id']:
                                        await member.send("해당 길드의 길드 마스터입니다.")
                                        await message.remove_reaction(payload.emoji, member)
                                        return
                                    else:
                                        await message.remove_reaction(payload.emoji, member)
                                        embed = discord.Embed(
                                            description=f"{member.mention}님, {guild_info['guild_name']} 길드 가입 대기 상태입니다.",
                                            color=discord.Color.yellow()
                                        )
                                        wait_msg = await member.send(embed=embed)
                                        guild_master = guild.get_member(guild_info['guild_leader_id'])
                                        embed = discord.Embed(
                                            title=f"{guild_role.name} 길드 가입 신청",
                                            description=f"{member.mention}님께서 {guild_info['guild_name']} 길드 가입을 신청하셨습니다.",
                                            color=discord.Color.yellow()
                                        )
                                        apply_msg = await guild_master.send(embed=embed)
                                        await apply_msg.add_reaction('⭕')
                                        await apply_msg.add_reaction('❌')
                                        
                                        bot.loop.create_task(wait_for_guild_master_decision(apply_msg, message, wait_msg, member, guild_info, guild))
                                        
                                        break
                
    
async def wait_for_guild_master_decision(apply_msg, guild_message, wait_msg, member, guild_info, guild):
    def check(reaction, user):
        return user != bot.user and str(reaction.emoji) in ['⭕', '❌'] and reaction.message.id == apply_msg.id
    guild_role = guild.get_role(guild_info['role_id'])
    reaction, user = await bot.wait_for('reaction_add', check=check)
    if str(reaction.emoji) == '⭕':
        await member.add_roles(guild_role)
        embed = discord.Embed(description=f"{member.mention}님의 {guild_info['guild_name']} 길드 가입 신청이 승인되었습니다.", color=discord.Color.green())
        await apply_msg.edit(embed=embed)
        embed = discord.Embed(description=f"{member.mention}님, {guild_info['guild_name']} 길드 가입 신청이 승인되었습니다.", color=discord.Color.green())
        await wait_msg.edit(embed=embed)
        guild_info['guild_members'].append(member.id)
        await update_guild_message(guild_message, guild_info)
        await save_guild_info(guild_info, guild)
    elif str(reaction.emoji) == '❌':
        embed = discord.Embed(description=f"{member.mention}님의 {guild_info['guild_name']} 길드 가입 신청이 거절되었습니다.", color=discord.Color.red())
        await apply_msg.edit(embed=embed)
        embed = discord.Embed(description=f"{member.mention}님, {guild_info['guild_name']} 길드 가입 신청이 거절되었습니다.", color=discord.Color.red())
        await wait_msg.edit(embed=embed)
        

async def save_guild_info(guild_info, guild):
    guilds = load_guilds()
    guild_id = str(guild.id)
    guild_name = guild_info['guild_name']
    guilds[guild_id][guild_name] = guild_info
    save_guilds(guilds)

@bot.command(name='랭킹시작', help='(운영진 전용)랭킹 업데이트를 시작합니다.\n사용법 : !랭킹시작')
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 랭킹시작(ctx):
    now = datetime.now(kst)
    configs = load_config()
    ranking_data = load_ranking()
    guild_id = str(ctx.guild.id)
    channel_id = ctx.channel.id
    if guild_id in configs:
        if 'ranking_channel_id' not in configs[guild_id]:
            if guild_id not in ranking_data:
                ranking_data[guild_id] = {}
            for member in ctx.guild.members:
                if not member.bot:
                    user_id = str(member.id)
                    if user_id not in ranking_data[guild_id]:
                        ranking_data[guild_id][user_id] = {"message_count": 0, "voice_time": 0}
            if guild_id in ranking_data:
                sorted_by_message_count = sorted(ranking_data[guild_id].items(), key=lambda x: x[1]['message_count'], reverse=True)
                sorted_by_voice_time = sorted(ranking_data[guild_id].items(), key=lambda x: x[1]['voice_time'], reverse=True)
            configs[guild_id]['ranking_channel_id'] = channel_id
            embed = discord.Embed(
                title=f"📊랭킹 현황({now.strftime('%Y년 %m월 %d일, %H:%M')})📊",
                description="랭킹은 매시 00분에 업데이트됩니다.",
                color=discord.Color.blue()
            )
            embed.add_field(name=f"⌨️ {ctx.guild.name} 키보드워리어 랭킹 ⌨️", value="\n".join([f"**{idx + 1}**. <@{user_id}>: {data['message_count']}개" for idx, (user_id, data) in enumerate(sorted_by_message_count[:5])]), inline=False)
            embed.add_field(name=f"👻 {ctx.guild.name} 지박령 랭킹 👻", value="\n".join([f"**{idx + 1}**. <@{user_id}>: {timedelta(seconds=int(data['voice_time']))}" for idx, (user_id, data) in enumerate(sorted_by_voice_time[:5])]), inline=False)
            view = PersistentView()
            msg = await ctx.send(embed=embed, view=view)
            configs[guild_id]['ranking_message_id'] = msg.id
            save_ranking(ranking_data)
            save_config(configs)
        else:
            msg = await ctx.send("이미 랭킹이 시작되었습니다.")
            await msg.delete(delay=5)

@bot.command(name='랭킹종료', help='(운영진 전용)랭킹 업데이트를 종료합니다.\n사용법 : !랭킹종료')
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 랭킹종료(ctx):
    ranking_data = load_ranking()
    configs = load_config()
    guild_id = str(ctx.guild.id)
    if guild_id in configs:
        configs[guild_id].pop('ranking_channel_id', None)
        configs[guild_id].pop('ranking_message_id', None)
        save_config(configs)
        await ctx.send("랭킹 업데이트를 종료합니다.")
    if guild_id in ranking_data:
        ranking_data.pop(guild_id)
        save_ranking(ranking_data)


@bot.command(name='추첨', help='(운영진 전용)특정 메시지에 특정 이모지를 남긴 사람들 중 당첨자를 추첨합니다.\n사용법 : !추첨 <메시지 ID> <이모지> <당첨 인원>')
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 추첨(ctx, message_id: int, emoji: str, number_of_winners: int):
    try:
        # 메시지 가져오기
        message = await ctx.channel.fetch_message(message_id)
        # 해당 이모지를 반응으로 남긴 사용자 목록 가져오기
        reaction = discord.utils.get(message.reactions, emoji=emoji)
        if reaction is None:
            await ctx.send(f"해당 메시지에 {emoji} 이모지가 없습니다.")
            return

        users = []
        async for user in reaction.users():
            if user != bot.user:
                users.append(user)

        # 추첨 인원이 전체 인원보다 많으면 오류 메시지 출력
        if number_of_winners > len(users):
            await ctx.send("추첨 인원이 반응한 인원보다 많습니다.")
            return

        # 추첨
        winners = random.sample(users, number_of_winners)

        # 당첨자 목록 출력
        winners_list = '\n'.join([winner.mention for winner in winners])
        message = await ctx.send("3초 후에 당첨자를 발표합니다.")
    
        for i in range(3, 0, -1):
            await asyncio.sleep(1)
            await message.delete()
            message = await ctx.send(f"{i}...")
        await asyncio.sleep(1)
        await message.delete()
        await ctx.send(
f"""
\U0001F973 축하드립니다! \U0001F973 
-----------------------------------
당첨 인원: {number_of_winners}명
당첨자: 
{winners_list}
-----------------------------------
"""
)

    except Exception as e:
        await ctx.send(f"오류 발생: {str(e)}")

@bot.command(name='청소', help='(운영진 전용)개수만큼 메시지를 삭제합니다.\n사용법 : !청소 <개수>')
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 청소(ctx, number: int):
    await ctx.channel.purge(limit=number + 1)

@bot.command(name='공지등록', help='(운영진 전용)반응으로 역할 자동 부여되는 공지사항 메시지 ID를 등록합니다.\n사용법 : !공지등록 <메시지 ID>')
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 공지등록(ctx, message_id: int):
    config = load_config()
    guild_id = str(ctx.guild.id)
    if guild_id not in config:
        config[guild_id] = {}
    if 'announcement_message_id' in config[guild_id]:
        if message_id == config[guild_id]['announcement_message_id']:
            await ctx.send("이미 해당 메시지 ID가 등록되어 있습니다.")
            return
        else:
            await ctx.send("기존 메시지 ID를 삭제하고 새로운 메시지 ID를 등록합니다.")
            config[guild_id].pop('announcement_message_id')

    config[guild_id]['announcement_message_id'] = message_id
    save_config(config)
    print(f"공지사항 메시지 ID가 {message_id}로 설정되었습니다.")

@bot.command(name='환영채널등록', help='(운영진 전용)환영인사 메시지가 올라오는 환영인사 채널 ID를 등록합니다.\n사용법 : !환영채널등록 <채널 ID>')
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 환영채널등록(ctx, channel: discord.TextChannel):
    config = load_config()
    guild_id = str(ctx.guild.id)
    if guild_id not in config:
        config[guild_id] = {}
    if 'welcome_channel_id' in config[guild_id]:
        if channel.id == config[guild_id]['welcome_channel_id']:
            await ctx.send("이미 해당 채널이 등록되어 있습니다.")
            return
        else:
            await ctx.send("기존 채널을 삭제하고 새로운 채널을 등록합니다.")
            config[guild_id].pop('welcome_channel_id')
            
    config[guild_id]['welcome_channel_id'] = channel.id
    save_config(config)
    print(f"환영 채널이 {channel.mention} 으로 설정되었습니다.")

@bot.command(name='도움말', help='사용 가능한 명령어 목록을 출력합니다.\n사용법 : !도움말')
async def 도움말(ctx):
    count = 0
    page = 1
    embed = discord.Embed(
        title=f"GM봇 도움말 {page}페이지",
        description="GM봇은 다양한 기능을 제공합니다.",
        color=discord.Color.blue()
    )
    user_roles = [role.name for role in ctx.author.roles]

    # 명령어를 이름 기준으로 정렬
    sorted_commands = sorted(bot.commands, key=lambda x: x.name)
    higher_commands = []
    for command in sorted_commands:
        if command.name == 'help':
            continue
        
        if command.help.startswith('(운영진 전용)'):
            if not any(role in user_roles for role in ['봇 관리자', '운영부', 'GM 관리자']):
                continue
            else:
                higher_commands.append(command)
                continue


        embed.add_field(name=f"**{bot.command_prefix}{command.name}**", value=f"```{command.help}```", inline=False)
        count += 1
        if count == 25:
            await ctx.author.send(embed=embed)
            embed.title = "GM봇 도움말 2페이지"
            embed.description = ""
            embed.clear_fields()
            count = 0
    if any(role in user_roles for role in ['봇 관리자', '운영부', 'GM 관리자']):
        embed.add_field(name="#__운영진 전용 명령어__#",value="⬇️⬇️⬇️아래부터는 운영진 전용 명령어입니다.⬇️⬇️⬇️" , inline=False)
        count += 1
        if count == 25:
            await ctx.author.send(embed=embed)
            embed.title = "GM봇 도움말 2페이지"
            embed.description = ""
            embed.clear_fields()
            count = 0
        for command in higher_commands:
            embed.add_field(name=f"**{bot.command_prefix}{command.name}**", value=f"```{command.help}```", inline=False)
            count += 1
            if count == 25:
                await ctx.author.send(embed=embed)
                embed.title = "GM봇 도움말 2페이지"
                embed.description = ""
                embed.clear_fields()
                count = 0
    await ctx.author.send(embed=embed)

@bot.command(name='길드도움말', help='길드 관련 명령어 목록을 출력합니다.\n사용법 : !길드도움말')
async def 길드도움말(ctx):
    embed = discord.Embed(
        title="GM봇 길드 도움말",
        description="길드 관련 명령어 목록입니다.",
        color=discord.Color.blue()
    )
    user_roles = [role.name for role in ctx.author.roles]
    higher_commands = []

    # 명령어를 이름 기준으로 정렬
    sorted_commands = sorted(bot.commands, key=lambda x: x.name)
    
    for command in sorted_commands:
        if not command.name.startswith('길드'):
            continue
        
        if command.name == 'help':
            continue
        
        if command.help.startswith('(운영진 전용)'):
            if not any(role in user_roles for role in ['봇 관리자', '운영부', 'GM 관리자']):
                continue
            else:
                higher_commands.append(command)
                continue

        embed.add_field(name=f"**{bot.command_prefix}{command.name}**", value=f"```{command.help}```", inline=False)
    if any(role in user_roles for role in ['봇 관리자', '운영부', 'GM 관리자']):
        embed.add_field(name="#__운영진 전용 명령어__#",value="⬇️⬇️⬇️아래부터는 운영진 전용 명령어입니다.⬇️⬇️⬇️" , inline=False)
        for command in higher_commands:
            embed.add_field(name=f"**{bot.command_prefix}{command.name}**", value=f"```{command.help}```", inline=False)
    await ctx.author.send(embed=embed)

@bot.command(name='길드카테고리등록', help='(운영진 전용)길드 카테고리를 등록합니다.\n사용법 : !길드카테고리등록 <카테고리 ID>')
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 길드카테고리등록(ctx, category: discord.CategoryChannel):
    guilds = load_guilds()
    guild_id = str(ctx.guild.id)
    if guild_id not in guilds:
        guilds[guild_id] = {}
    if 'category_id' in guilds[guild_id]:
        if category.id == guilds[guild_id]['category_id']:
            await ctx.send("이미 해당 카테고리가 등록되어 있습니다.")
            return
        else:
            await ctx.send("기존 카테고리를 삭제하고 새로운 카테고리를 등록합니다.")
            guilds[guild_id].pop('category_id')

    guilds[guild_id]['category_id'] = category.id
    save_guilds(guilds)
    await ctx.send(f"길드 카테고리가 {category.name}으로 설정되었습니다.")

@bot.command(name='길드목록채널등록', help='(운영진 전용)길드 목록 채널을 등록합니다.\n사용법 : !길드목록채널등록 <채널 ID>')
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 길드목록채널등록(ctx, channel: discord.TextChannel):
    guilds = load_guilds()
    guild_id = str(ctx.guild.id)
    if guild_id not in guilds:
        guilds[guild_id] = {}
    if 'list_channel_id' in guilds[guild_id]:
        if channel.id == guilds[guild_id]['list_channel_id']:
            await ctx.send("이미 해당 채널이 등록되어 있습니다.")
            return
        else:
            await ctx.send("기존 채널을 삭제하고 새로운 채널을 등록합니다.")
            guilds[guild_id].pop('list_channel_id')
    
    guilds[guild_id]['list_channel_id'] = channel.id
    save_guilds(guilds)
    await ctx.send(f"길드 목록 채널이 {channel.name}으로 설정되었습니다.")

@bot.command(name='길드생성', help='길드를 생성합니다.\n길드명은 띄어쓰기가 불가능합니다.\n길드 설명은 띄어쓰기 및 줄바꿈(쉬프트+엔터)이 가능합니다.\n사용법 : !길드생성 <길드명> <길드 설명>')
async def 길드생성(ctx, name: str, *, description: str):
    guilds = load_guilds()
    guild_id = str(ctx.guild.id)
    created_role = None
    created_channel = None
    created_message = None

    try:
        if guild_id not in guilds:
            guilds[guild_id] = {}

        category_id = guilds[guild_id].get('category_id')
        if category_id is None:
            message = await ctx.send("길드 카테고리가 설정되지 않았습니다.")
            await asyncio.sleep(2)
            await message.delete()
            return

        category = ctx.guild.get_channel(category_id)
        if category is None:
            message = await ctx.send("길드 카테고리를 찾을 수 없습니다.")
            await asyncio.sleep(2)
            await message.delete()
            return

        author_avatar_url = ctx.author.avatar.url if ctx.author.avatar else ctx.author.default_avatar.url

        embed = discord.Embed(
            title=f"{name} 길드",
            color=discord.Color.blue()
        )
        embed.set_author(name=ctx.author.display_name, icon_url=author_avatar_url)
        embed.add_field(name="길드 설명", value=f">>> {description}", inline=False)
        embed.add_field(name="길드 마스터", value=ctx.author.mention, inline=False)
        embed.add_field(name="길드원", value="", inline=False)
        embed.add_field(name="길드 가입", value="📝 이모지를 누르면 길드 마스터에게 길드 가입 신청 메시지가 전송됩니다. 이후, 길드 마스터의 승인 하에 길드에 가입하실 수 있습니다.", inline=False)
        embed.set_thumbnail(url=author_avatar_url)

        guild_role = await ctx.guild.create_role(name=f"{name}(길드)")

        overwrites = {
            ctx.guild.default_role: discord.PermissionOverwrite(view_channel=False),
            guild_role: discord.PermissionOverwrite(view_channel=True)
        }

        list_channel = bot.get_channel(guilds[guild_id].get('list_channel_id'))

        guild = await ctx.guild.create_voice_channel(name, category=category, overwrites=overwrites)

        await ctx.author.add_roles(guild_role)
        button = Button(style=discord.ButtonStyle.primary, label="가입신청", custom_id=f"apply_{guild_id}_{guild_role.id}")
        created_message = await list_channel.send(embed=embed)
        await ctx.send(f"{name} 길드가 생성되었습니다.")

        guilds[guild_id][name] = {
            'guild_name': name,
            'guild_leader_id': ctx.author.id,
            'role_id': guild_role.id,
            'channel_id': guild.id,
            'message_id': created_message.id
        }
        await created_message.add_reaction('📝')
        save_guilds(guilds)

    except Exception as e:
        if created_role:
            await created_role.delete()
        if created_channel:
            await created_channel.delete()
        if created_message:
            await created_message.delete()
        await ctx.author.send(f"오류 발생: {str(e)}. 길드 생성에 실패했습니다.")

@bot.command(name='길드삭제', help='(운영진 전용)길드를 삭제합니다.\n사용법 : !길드삭제 <길드명>')
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 길드삭제(ctx, name: str):
    guilds = load_guilds()
    guild_id = str(ctx.guild.id)
    if guild_id not in guilds or name not in guilds[guild_id]:
        await ctx.author.send("해당 길드를 찾을 수 없습니다.")
        return

    guild_info = guilds[guild_id][name]
    guild = ctx.guild.get_channel(guild_info['channel_id'])
    if guild is not None:
        await guild.delete()

    role = ctx.guild.get_role(guild_info['role_id'])
    if role is not None:
        await role.delete()

    list_channel = ctx.guild.get_channel(guilds[guild_id].get('list_channel_id'))
    message = await list_channel.fetch_message(guild_info['message_id'])
    await message.delete()

    del guilds[guild_id][name]
    save_guilds(guilds)
    await ctx.author.send(f"{name} 길드가 삭제되었습니다.")

@bot.command(name='길드탈퇴', help='길드를 탈퇴합니다.\n사용법 : !길드탈퇴 <길드명>')
async def 길드탈퇴(ctx, name: str):
    guilds = load_guilds()
    guild_id = str(ctx.guild.id)
    if guild_id not in guilds or name not in guilds[guild_id]:
        await ctx.author.send("해당 길드를 찾을 수 없습니다.")
        return

    guild_info = guilds[guild_id][name]

    if ctx.author.id == guild_info['guild_leader_id']:
        await ctx.author.send("해당 길드의 길드 마스터는 탈퇴할 수 없습니다.")
        return
    
    guild_role = ctx.guild.get_role(guild_info['role_id'])
    if guild_role is not None:
        await ctx.author.remove_roles(guild_role)

    list_channel = ctx.guild.get_channel(guilds[guild_id].get('list_channel_id'))
    message = await list_channel.fetch_message(guild_info['message_id'])
    if ctx.author.id in guild_info['guild_members']:
        guild_info['guild_members'].remove(ctx.author.id)
        await update_guild_message(message, guild_info)
        save_guilds(guilds)
    await ctx.author.send(f"{name} 길드에서 탈퇴되었습니다.")

@bot.command(name='길드멤버퇴출', help='길드 멤버를 퇴출합니다.(길드 마스터 전용)\n사용법 : !길드멤버퇴출 <길드명> <사용자ID>')
async def 길드멤버퇴출(ctx, name: str, member_id: int):
    guilds = load_guilds()
    guild_id = str(ctx.guild.id)
    
    if guild_id not in guilds or name not in guilds[guild_id]:
        await ctx.author.send("해당 길드를 찾을 수 없습니다.")
        return

    guild_info = guilds[guild_id][name]

    if ctx.author.id != guild_info['guild_leader_id']:
        await ctx.author.send("해당 길드의 길드 마스터만 길드원을 퇴출할 수 있습니다.")
        return
    
    member = ctx.guild.get_member(member_id)
    if member is None:
        await ctx.author.send("사용자를 찾을 수 없습니다.")
        return
    
    guild_role = ctx.guild.get_role(guild_info['role_id'])
    if guild_role is None:
        await ctx.author.send("길드 역할을 찾을 수 없습니다.")
        return

    if member.id == guild_info['guild_leader_id']:
        await ctx.author.send("길드 마스터는 퇴출할 수 없습니다.")
        return

    list_channel = ctx.guild.get_channel(guilds[guild_id].get('list_channel_id'))
    message = await list_channel.fetch_message(guild_info['message_id'])

    if member.id in guild_info['guild_members']:
        guild_info['guild_members'].remove(member.id)
        await update_guild_message(message, guild_info)
        save_guilds(guilds)
        await member.remove_roles(guild_role)
        await member.send(f"{name} 길드에서 퇴출되었습니다.")
    else:
        await ctx.send("해당 사용자는 길드에 가입되어 있지 않습니다.")

@bot.command(name='길드명변경', help='길드명을 변경합니다.(길드 마스터 전용)\n사용법 : !길드명변경 <기존 길드명> <새 길드명>')
async def 길드명변경(ctx, old_name: str, new_name: str):
    guilds = load_guilds()
    guild_id = str(ctx.guild.id)
    if guild_id not in guilds or old_name not in guilds[guild_id]:
        await ctx.author.send("해당 길드를 찾을 수 없습니다.")
        return

    guild_info = guilds[guild_id][old_name]
    if ctx.author.id != guild_info['guild_leader_id']:
        await ctx.author.send("해당 길드의 길드 마스터만 길드명을 변경할 수 있습니다.")
        return
    
    guild_role = ctx.guild.get_role(guild_info['role_id'])
    if guild_role is not None:
        await guild_role.edit(name=f"{new_name}(길드)")

    list_channel = ctx.guild.get_channel(guilds[guild_id].get('list_channel_id'))
    message = await list_channel.fetch_message(guild_info['message_id'])
    embed = message.embeds[0]
    embed.title = f"{new_name} 길드"
    await message.edit(embed=embed)

    guilds[guild_id][new_name] = guilds[guild_id].pop(old_name)
    save_guilds(guilds)
    await ctx.author.send(f"{old_name} 길드명이 {new_name}으로 변경되었습니다.")

async def update_guild_message(message, guild_info):
    member_mentions = []
    for member_id in guild_info['guild_members']:
        member = message.guild.get_member(member_id)
        if member:
            member_mentions.append(member.mention)
    updated_members = ', '.join(member_mentions)
    embed = message.embeds[0]
    embed.set_field_at(2, name="길드원", value=updated_members, inline=False)
    await message.edit(embed=embed)

@bot.command(name="출석채널등록", help="(운영진 전용)회원 출석 채널을 등록합니다.\n사용법 : !출석채널등록 <채널 ID>")
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 출석채널등록(ctx, channel: discord.TextChannel):
    config = load_config()
    guild_id = str(ctx.guild.id)
    if guild_id not in config:
        config[guild_id] = {}
    if 'attendance_channel_id' in config[guild_id]:
        if channel.id == config[guild_id]['attendance_channel_id']:
            await ctx.send("이미 해당 채널이 등록되어 있습니다.")
            return
        else:
            await ctx.send("기존 채널을 삭제하고 새로운 채널을 등록합니다.")
            config[guild_id].pop('attendance_channel_id')
    
    config[guild_id]['attendance_channel_id'] = channel.id
    save_config(config)
    await ctx.send(f"출석 채널이 {channel.name}으로 설정되었습니다.")

@bot.command(name="출석시작", help="(운영진 전용)출석 체크를 시작합니다.\n사용법 : !출석시작")
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 출석시작(ctx):
    configs = load_config()
    guild_id = str(ctx.guild.id)
    today_date = datetime.now(kst).strftime("%y.%m.%d")

    file_name = f"{ctx.guild.name}_출석부.xlsx"
    if guild_id in configs:
        if 'is_checking_attendance' in configs[guild_id] and configs[guild_id]['is_checking_attendance']:
            await ctx.send("이미 출석이 진행중입니다.")
            return
    members = ctx.guild.members
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "출석부"
    date_col = 1
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False),start=2):
        if row[date_col-1].value == today_date:
            await ctx.send("당일 출석은 이미 종료되었습니다.")
            return
    sheet.append(["날짜", "이름", "출결"])
    datas = []
    for member in members:
        if not member.bot:
            match = re.search(r'\[.*?\]\s*(.*)', member.display_name)
            if match:
                clean_name = match.group(1).replace(" ", "")
            else:
                clean_name = member.display_name.replace(" ", "")
            datas.append([today_date, clean_name, "결석"])
    datas.sort(key=lambda x: x[1])
    for data in datas:
        sheet.append(data)
    sheet.append(["------------", "------------", "------------"])
    workbook.save(file_name)

    datas.clear()

    buttons = []
    for member in members:
        if not member.bot:
            match = re.search(r'\[.*?\]\s*(.*)', member.display_name)
            if match:
                clean_name = match.group(1).replace(" ", "")
            else:
                clean_name = member.display_name.replace(" ", "")
            datas.append([clean_name,f"onTime_{member.id}"])
    
    datas.sort(key=lambda x: x[0])
    for data in datas:
        button = discord.ui.Button(style=discord.ButtonStyle.green, label=data[0], custom_id=data[1])
        buttons.append(button)

    view = discord.ui.View()
    for button in buttons:
        view.add_item(button)
    

    if guild_id in configs:
        if 'attendance_button_id' in configs[guild_id]:
            button_id = configs[guild_id]['attendance_button_id']
            try:
                button_message = await ctx.channel.fetch_message(button_id)
                await button_message.delete()
            except Exception as e:
                print(f"메시지를 찾을 수 없습니다. {str(e)}")
    
    view_message = await ctx.send(f"{today_date} 출석 체크를 시작합니다.", view=view)

    if guild_id in configs:
        configs[guild_id]['attendance_button_id'] = view_message.id
        configs[guild_id]['is_checking_attendance'] = True

    embed = discord.Embed(
        title=f"{today_date} 출석 현황",
        description="출석 진행중",
        color=discord.Color.green()
    )
    embed.add_field(name="출석", value="", inline=False)
    embed.add_field(name="지각", value="", inline=False)
    if guild_id in configs:
        if 'attendance_channel_id' in configs[guild_id]:
            attendance_channel_id = configs[guild_id]['attendance_channel_id']
            attendance_channel = bot.get_channel(attendance_channel_id)
            if attendance_channel:
                message = await attendance_channel.send(embed=embed)
                configs[guild_id]['attendance_message_id'] = message.id
    
    save_config(configs)

@bot.command(name="지각시작", help="(운영진 전용)지각 체크를 시작합니다.\n사용법 : !지각시작")
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 지각시작(ctx):
    configs = load_config()
    guild_id = str(ctx.guild.id)
    if guild_id in configs:
        if 'is_checking_attendance' in configs[guild_id] and not configs[guild_id]['is_checking_attendance']:
            await ctx.send("진행중인 출석이 없습니다.")
            return
    members = ctx.guild.members
    datas = []
    buttons = []
    for member in members:
        if not member.bot:
            match = re.search(r'\[.*?\]\s*(.*)', member.display_name)
            if match:
                clean_name = match.group(1).replace(" ", "")
            else:
                clean_name = member.display_name.replace(" ", "")
            datas.append([clean_name,f"onLate_{member.id}"])
    
    datas.sort(key=lambda x: x[0])
    for data in datas:
        button = discord.ui.Button(style=discord.ButtonStyle.red, label=data[0], custom_id=data[1])
        buttons.append(button)

    view = discord.ui.View()
    for button in buttons:
        view.add_item(button)
    
    configs = load_config()
    guild_id = str(ctx.guild.id)

    if guild_id in configs:
        if 'attendance_button_id' in configs[guild_id]:
            button_id = configs[guild_id]['attendance_button_id']
            try:
                button_message = await ctx.channel.fetch_message(button_id)
                await button_message.delete()
            except Exception as e:
                print(f"메시지를 찾을 수 없습니다. {str(e)}")
    today_date = datetime.now(kst).strftime("%y.%m.%d")
    view_message = await ctx.send(f"{today_date} 지각 체크를 시작합니다.", view=view)
    configs[guild_id]['attendance_button_id'] = view_message.id
    save_config(configs)

@bot.command(name="출석종료", help="(운영진 전용)출석부를 종료하고 엑셀 파일을 저장합니다.\n사용법 : !출석종료")
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 출석종료(ctx):
    configs = load_config()
    guild_id = str(ctx.guild.id)
    message = None
    if guild_id in configs:
        configs[guild_id]['is_checking_attendance'] = False
        if 'attendance_channel_id' in configs[guild_id]:
            attendance_channel_id = configs[guild_id]['attendance_channel_id']
            attendance_channel = bot.get_channel(attendance_channel_id)
            if attendance_channel:
                message_id = configs[guild_id]['attendance_message_id']
                message = await attendance_channel.fetch_message(message_id)
        if 'attendance_button_id' in configs[guild_id]:
            button_id = configs[guild_id]['attendance_button_id']
            try:
                button_message = await ctx.channel.fetch_message(button_id)
                await button_message.delete()
            except Exception as e:
                print(f"메시지를 찾을 수 없습니다. {str(e)}")
    save_config(configs)
    try:
        if message:
            embed = message.embeds[0]
            embed.description = "출석 종료"
            embed.color = discord.Color.red()
            await message.edit(embed=embed)
    except Exception as e:
        await ctx.send(f"오류 발생: {str(e)}")
    today_date = datetime.now(kst).strftime("%y.%m.%d")
    await 출석부(ctx, today_date)

@bot.command(name="출석예약", help="출석을 예약합니다.\n사용법 : !출석예약 <출석시작시간 hh:mm> <지각시작시간 hh:mm> <출석종료시간 hh:mm>")
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 출석예약(ctx, start_time: str, late_time: str, end_time: str):
    try:
        datetime.strptime(start_time, "%H:%M")
        datetime.strptime(late_time, "%H:%M")
        datetime.strptime(end_time, "%H:%M")
    except ValueError:
        await ctx.author.send("시간 형식이 잘못되었습니다. 시간 형식은 hh:mm입니다.")
        return

    await ctx.author.send(f"출석 예약이 완료되었습니다.\n출석 시간: {start_time}\n지각 시간: {late_time}\n출석 종료 시간: {end_time}")
    bot.loop.create_task(schedule_attendance(ctx, start_time, late_time, end_time))

async def schedule_attendance(ctx, start_time, late_time, end_time):
    now = datetime.now(kst)
    print(f"현재 시간: {now}")

    start_time_obj = datetime.strptime(start_time, "%H:%M").replace(year=now.year, month=now.month, day=now.day)
    start_time_obj = kst.localize(start_time_obj)
    delay = (start_time_obj - now).total_seconds()
    print(f"출석 시작까지 대기 시간: {delay}")
    if delay > 0:
        await asyncio.sleep(delay)
        await 출석시작(ctx)
    
    now = datetime.now(kst)
    late_time_obj = datetime.strptime(late_time, "%H:%M").replace(year=now.year, month=now.month, day=now.day)
    late_time_obj = kst.localize(late_time_obj)
    delay = (late_time_obj - now).total_seconds()
    print(f"지각 시작까지 대기 시간: {delay}")
    if delay > 0:
        await asyncio.sleep(delay)
        await 지각시작(ctx)
    
    now = datetime.now(kst)
    end_time_obj = datetime.strptime(end_time, "%H:%M").replace(year=now.year, month=now.month, day=now.day)
    end_time_obj = kst.localize(end_time_obj)
    delay = (end_time_obj - now).total_seconds()
    print(f"출석 종료까지 대기 시간: {delay}")
    if delay > 0:
        await asyncio.sleep(delay)
        await 출석종료(ctx)

@bot.event
async def on_interaction(interaction: discord.Interaction):
    if interaction.type == discord.InteractionType.component:
        if interaction.custom_id.startswith('onTime_') or interaction.custom_id.startswith('onLate_'):
            configs = load_config()
            guild_id = str(interaction.guild_id)
            if guild_id in configs:
                if 'is_checking_attendance' in configs[guild_id]:
                    if not configs[guild_id]['is_checking_attendance']:
                        await interaction.response.send_message("출석이 종료되었습니다.", ephemeral=True)
                        return
            member_id = int(interaction.custom_id.split('_')[1])
            member = interaction.guild.get_member(member_id)
            await interaction.response.defer()
            if member:
                if interaction.custom_id.startswith('onTime_'):
                    embed = discord.Embed(
                        title="출석",
                        description=f"출석 버튼을 눌러 출석하세요.",
                        color=discord.Color.green()
                    )
                    button = discord.ui.Button(style=discord.ButtonStyle.green, label="출석", custom_id=f"attend_{guild_id}_{member_id}")
                else:
                    embed = discord.Embed(
                        title="지각",
                        description=f"지각 버튼을 눌러 지각 처리하세요.",
                        color=discord.Color.red()
                    )
                    button = discord.ui.Button(style=discord.ButtonStyle.red, label="지각", custom_id=f"late_{guild_id}_{member_id}")
                View = discord.ui.View()
                View.add_item(button)               
                await member.send(embed=embed, view=View)
        elif interaction.custom_id.startswith('attend_') or interaction.custom_id.startswith('late_'):
            guild_id = interaction.custom_id.split('_')[1]
            guild = bot.get_guild(int(guild_id))
            member_id = int(interaction.custom_id.split('_')[2])
            await interaction.message.delete()
            await interaction.response.defer()
            if interaction.custom_id.startswith('attend_'):
                await attend(guild, member_id)
            else:
                await attend(guild, member_id, True)

async def attend(guild, member_id, isLate=False):
    today_date = datetime.now(kst).strftime("%y.%m.%d")
    file_name = f"{guild.name}_출석부.xlsx"
    try:
        member = guild.get_member(member_id)
        member_name = guild.get_member(member_id).display_name
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active

        match = re.search(r'\[.*?\]\s*(.*)', member_name)
        if match:
            member_name = match.group(1).replace(" ", "")
        else:
            member_name = member_name.replace(" ", "")

        date_col = 1
        name_col = 2
        attendance_col = 3

        configs = load_config()
        guild_id = str(guild.id)
        message = None
        if guild_id in configs:
            if 'attendance_channel_id' in configs[guild_id]:
                attendance_channel_id = configs[guild_id]['attendance_channel_id']
                attendance_channel = bot.get_channel(attendance_channel_id)
                if attendance_channel:
                    message_id = configs[guild_id]['attendance_message_id']
                    message = await attendance_channel.fetch_message(message_id)

        name_found = False
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False),start=2):
            if row[name_col-1].value == member_name and row[date_col-1].value == today_date:
                name_found = True
                if isLate:
                    sheet.cell(row = idx, column = attendance_col, value = "지각")
                    embed = discord.Embed(
                        title=f"{today_date}",
                        description="지각",
                        color=discord.Color.red()
                    )
                else:
                    sheet.cell(row = idx, column = attendance_col, value = "출석")
                    embed = discord.Embed(
                        title=f"{today_date}",
                        description="출석",
                        color=discord.Color.green()
                    )
                await member.send(embed=embed)
                workbook.save(file_name)
                
                if message:
                    await update_attendance_message(message, member, isLate)
                else:
                    embed = discord.Embed(
                        title=f"오류",
                        description="출석 코드가 일치하지 않습니다.",
                        color=discord.Color.red()
                    )
                    await member.send(embed=embed)
                break
        if not name_found:
            embed = discord.Embed(
                title=f"오류",
                description="이름을 찾을 수 없습니다.",
                color=discord.Color.red()
            )
            await member.send(embed=embed)
    except FileNotFoundError:
        await member.send("출석부 파일을 찾을 수 없습니다.")
    except Exception as e:
        await member.send(f"오류 발생: {str(e)}")


async def update_attendance_message(message, member, isLate=False):
    embed = message.embeds[0]
    if isLate:
        late_field = embed.fields[1]
        late_field.value += f"{member.mention}\n"
        embed.set_field_at(1, name="지각", value=late_field.value, inline=False)
    else:
        attendance_field = embed.fields[0]
        attendance_field.value += f"{member.mention}\n"
        embed.set_field_at(0, name="출석", value=attendance_field.value, inline=False)
    await message.edit(embed=embed)

@bot.command(name="출석부", help="(운영진 전용)출석부를 출력합니다.\n사용법 :\n해당 날짜 출석부: !출석부 <날짜(예: 24.08.08)>\n전체 출석부: !출석부")
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 출석부(ctx, date: str = None):
    if date:
        try:
            parsed_date = datetime.strptime(date, "%y.%m.%d")
            formatted_date = parsed_date.strftime("%y.%m.%d")
        except ValueError:
            await ctx.send("날짜 형식이 잘못되었습니다. (예: 24.08.08)")
            return

    file_name = f"{ctx.guild.name}_출석부.xlsx"
    try:
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        await ctx.send("출석부 파일을 찾을 수 없습니다.")
        return
    sheet = workbook.active
    date_col = 1
    name_col = 2
    attendance_col = 3
    dates = []
    names = []
    attendances = []
    if date:
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False),start=2):
            if row[date_col-1].value == formatted_date:
                names.append(row[name_col-1].value)
                attendances.append(row[attendance_col-1].value)
    else:
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False),start=2):
            dates.append(row[date_col-1].value)
            names.append(row[name_col-1].value)
            attendances.append(row[attendance_col-1].value)
    if date:
        embed = discord.Embed(
            title=f"{formatted_date} 출석부",
            color=discord.Color.blue()
        )
        embed.add_field(name="이름", value="\n".join(names), inline=True)
        embed.add_field(name="출결", value="\n".join(attendances), inline=True)
    else:
        embed = discord.Embed(
            title="출석부",
            color=discord.Color.blue()
        )
        embed.add_field(name="날짜", value="\n".join(dates), inline=True)
        embed.add_field(name="이름", value="\n".join(names), inline=True)
        embed.add_field(name="출결", value="\n".join(attendances), inline=True)
    await ctx.send(embed=embed)

@bot.command(name="공결", help="(운영진 전용)공결 처리를 합니다.\n사용법 : !공결 <날짜(예: 24.08.08)> <이름>")
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 공결(ctx, date: str, name: str):
    await update_attendance(ctx.guild, ctx, "공결", date, name)

@bot.command(name="결석", help="(운영진 전용)결석 처리를 합니다.\n사용법 : !결석 <날짜(예: 24.08.08)> <이름>")
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 결석(ctx, date: str, name: str):
    await update_attendance(ctx.guild, ctx, "결석", date, name)

@bot.command(name="출석", help="(운영진 전용)출석 처리를 합니다.\n사용법 : !출석 <날짜(예: 24.08.08)> <이름>")
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 출석(ctx, date: str, name: str):
    await update_attendance(ctx.guild, ctx, "출석", date, name)

@bot.command(name="지각", help="(운영진 전용)지각 처리를 합니다.\n사용법 : !지각 <날짜(예: 24.08.08)> <이름>")
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 지각(ctx, date: str, name: str):
    await update_attendance(ctx.guild, ctx, "지각", date, name)

async def update_attendance(guild, ctx, state, date, name):
    file_name = f"{guild.name}_출석부.xlsx"
    try:
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        await ctx.send("출석부 파일을 찾을 수 없습니다.")
        return
    sheet = workbook.active
    date_col = 1
    name_col = 2
    attendance_col = 3
    try:
        parsed_date = datetime.strptime(date, "%y.%m.%d")
        formatted_date = parsed_date.strftime("%y.%m.%d")
    except ValueError:
        await ctx.send("날짜 형식이 잘못되었습니다. (예: 24.08.08)")
        return
    name_found = False
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False),start=2):
        if row[date_col-1].value == formatted_date and row[name_col-1].value == name:
            name_found = True
            sheet.cell(row = idx, column = attendance_col, value = state)
            await ctx.send(f"{formatted_date} {name} {state} 처리되었습니다.")
            break
    if not name_found:
        await ctx.send("이름을 찾을 수 없습니다.")
    workbook.save(file_name)

@bot.command(name="출석부다운", help="(운영진 전용)출석부 엑셀 파일을 다운로드합니다.\n사용법 : !출석부다운")
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 출석부다운(ctx):
    file_name = f"{ctx.guild.name}_출석부.xlsx"
    try:
        await ctx.send(file=discord.File(file_name))
    except FileNotFoundError:
        await ctx.send("출석부 파일을 찾을 수 없습니다.")

@bot.command(name="링크등록", help="(운영진 전용)동아리 공식 SNS, 회계장부, 회칙 등 링크를 등록합니다.\n사용법 : !링크등록 <링크 이름> <링크>")
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 링크등록(ctx, link_name: str, link: str):
    configs = load_config()
    guild_id = str(ctx.guild.id)
    if guild_id not in configs:
        configs[guild_id] = {}
    if 'links' not in configs[guild_id]:
        configs[guild_id]['links'] = {}
    configs[guild_id]['links'][link_name] = link
    save_config(configs)
    await ctx.author.send("링크가 등록되었습니다.")

@bot.command(name="바로가기", help="바로가기 링크를 출력합니다.\n사용법 : !바로가기")
async def 바로가기(ctx):
    configs = load_config()
    guild_id = str(ctx.guild.id)
    if guild_id in configs:
        embed = discord.Embed(
            title=f"{ctx.guild.name} 바로가기",
            description=f"{ctx.guild.name} 공식 SNS, 회계장부 등 공개 자료 바로가기입니다.",
            color=discord.Color.blue()
        )
        view = discord.ui.View()
        for link_name, link in configs[guild_id]['links'].items():
            embed.add_field(name=link_name, value=link, inline=False)
            button = discord.ui.Button(style=discord.ButtonStyle.link, label=link_name, url=link)
            view.add_item(button)
            
        await ctx.send(embed=embed, view=view)

@bot.command(name="링크삭제", help="(운영진 전용)등록된 링크를 삭제합니다.\n사용법 : !링크삭제")
@commands.has_any_role('봇 관리자', '운영부', 'GM 관리자')
async def 링크삭제(ctx, link_name: str):
    configs = load_config()
    guild_id = str(ctx.guild.id)
    if guild_id in configs:
        if 'links' in configs[guild_id]:
            if link_name in configs[guild_id]['links']:
                configs[guild_id]['links'].pop(link_name)
                save_config(configs)
                await ctx.author.send("링크가 삭제되었습니다.")
                return


# 봇 실행
bot.run()